import pickle
from datetime import datetime
import cv2
import imutils
import numpy
import os
import requests
from imutils.video import FPS
from imutils.video import WebcamVideoStream
from openpyxl import load_workbook, Workbook
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders


subject_name = input("Enter Subject Name: ")
teacher_email = input("Enter Teacher Email: ")

my_file_name = f'student-attendance-{subject_name}.xlsx'
db_file_name = f"db/datafile-{subject_name}.pickle"

gmail_user = "raspberrycar2017@gmail.com"
gmail_pwd = "hfmrwbwuhxdmxmwa"

student_chk = []
root, STUDENT_LIST, files = next(os.walk('datasets'), ([], [], []))
present_students = []  # has strings with present

day = int(datetime.now().strftime("%d"))
month = str(datetime.now().strftime("%b"))
today_full_date = str(datetime.now().strftime('%d-%b-%Y'))

if not os.path.exists(db_file_name):
    pickle.dump({}, open(db_file_name, 'wb'))

if not os.path.exists(my_file_name):
    _wb = Workbook()
    _ws = _wb["Sheet"]
    _ws.cell(column=1, row=1, value="Name")
    for j in range(1, 32):
        _ws.cell(column=j + 1, row=1, value=f"{j}/{month}/22")
    _ws.cell(column=33, row=1, value="Percentage")
    _wb.save(my_file_name)
    _wb.close()

_names_database = pickle.load(open(db_file_name, 'rb'))

new_dict = {}
for _, itm in enumerate(STUDENT_LIST):
    if str(itm) in _names_database:
        new_dict[itm] = _names_database[itm]
    else:
        arr1 = []
        for i in range(0, 32):  # 32 is percent
            arr1.append("")
        new_dict[itm] = arr1

_names_database = new_dict

mywb = load_workbook(filename=my_file_name)
myws = mywb["Sheet"]


def getCountOfPresent(mylist):
    times = 0
    for occurrence in mylist:
        if occurrence == 'Present':
            times += 1
    return times


def saveExcelFile():
    myws.cell(column=1, row=1, value="Name")
    for cnt, item in enumerate(_names_database):
        myws.cell(column=1, row=cnt + 2, value=item)  # name
        for _i in range(0, 31):
            myws.cell(column=_i + 2, row=cnt + 2, value=_names_database[item][_i])

        present_count = getCountOfPresent(_names_database[item])
        myws.cell(column=33, row=cnt + 2, value=str(round((present_count / 31) * 100, 2)) + ' %')

    mywb.save(filename=my_file_name)
    pickle.dump(_names_database, open(db_file_name, 'wb'))


def saveAttendance(name, attn):
    _names_database[name][day - 1] = attn


def mail(to, subject, text, attach):
    msg = MIMEMultipart()

    msg['From'] = gmail_user
    msg['To'] = to
    msg['Subject'] = subject

    msg.attach(MIMEText(text))

    part = MIMEBase('application', 'octet-stream')
    part.set_payload(open(attach, 'rb').read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition',
                    'attachment; filename="%s"' % os.path.basename(attach))
    msg.attach(part)

    mail_server = smtplib.SMTP("smtp.gmail.com", 587)
    mail_server.ehlo()
    mail_server.starttls()
    mail_server.ehlo()
    mail_server.login(gmail_user, gmail_pwd)
    mail_server.sendmail(gmail_user, to, msg.as_string())
    # Should be mail_server.quit(), but that crashes...
    mail_server.close()


def main():
    while True:
        im = webcam.read()
        im = imutils.resize(im, width=400)
        gray = cv2.cvtColor(im, cv2.COLOR_BGR2GRAY)
        faces = face_cascade.detectMultiScale(gray, 1.3, 5)
        for (x, y, w, h) in faces:
            face = gray[y:y + h, x:x + w]
            face_resize = cv2.resize(face, (width, height))
            prediction = model.predict(face_resize)
            cv2.rectangle(im, (x, y), (x + w, y + h), (0, 255, 0), 3)

            if prediction[1] >= 60:
                nam = names[prediction[0]]
                if nam not in student_chk:
                    student_chk.append(nam)
                    STUDENT_LIST.remove(nam)
                    student_present_str = f"{names[prediction[0]]} - Present on {today_full_date}"
                    present_students.append(student_present_str)
                    saveAttendance(nam, 'Present')
                    print(student_present_str)

                cv2.rectangle(im, (x, y), (x + w, y + h), (255, 0, 0), 3)
                cv2.putText(im, '%s - %.0f' % (names[prediction[0]], prediction[1]), (x - 10, y - 10),
                            cv2.FONT_HERSHEY_PLAIN, 1, (0, 255, 0))
            else:
                cv2.rectangle(im, (x, y), (x + w, y + h), (0, 0, 255), 3)
                cv2.putText(im, 'not recognized', (x - 10, y - 10), cv2.FONT_HERSHEY_PLAIN, 1, (0, 255, 0))

        cv2.imshow('OpenCV', im)
        key = cv2.waitKey(10)
        if key == 27:
            fps.stop()
            cv2.destroyAllWindows()
            break
        if key % 256 == ord('m'):
            print("sending data to server... please wait..")
            arr = []

            for index, item in enumerate(present_students):
                arr.append(item)

            for index, item in enumerate(STUDENT_LIST):
                abs_str = f"{item} - Absent on {today_full_date}"
                arr.append(abs_str)
                saveAttendance(item, 'Absent')

            saveExcelFile()

            requests.post(url='https://student-attendance-system-iot.herokuapp.com/api.php?auth=Zbz9yPzE5',
                          json={"students": arr})
            mail(teacher_email, f'Student Attendance Report for {subject_name}',
                 f'Student Attendance Report for {subject_name}',
                 my_file_name)
            print("data sent!")
        fps.update()


haar_file = 'haarcascade_frontalface_default.xml'
datasets = 'datasets'

print('Training...')
# Create a list of images and a list of corresponding names
(images, labels, names, id) = ([], [], {}, 0)
for (subdirs, dirs, files) in os.walk(datasets):
    for subdir in dirs:
        names[id] = subdir
        subjectpath = os.path.join(datasets, subdir)
        for filename in os.listdir(subjectpath):
            path = subjectpath + '/' + filename
            label = id
            images.append(cv2.imread(path, 0))
            labels.append(int(label))
        id += 1
(width, height) = (130, 100)

# Create a Numpy array from the two lists above
(images, labels) = [numpy.array(lis) for lis in [images, labels]]

model = cv2.face.LBPHFaceRecognizer_create()
model.train(images, labels)

# use LBPHFace recognizer on camera frame
face_cascade = cv2.CascadeClassifier(haar_file)
webcam = WebcamVideoStream(src=0).start()

fps = FPS().start()
main()
